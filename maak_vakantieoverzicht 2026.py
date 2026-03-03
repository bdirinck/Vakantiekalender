"""
De uitvoeringsomgeving werkt momenteel niet, dus ik geef je het aangepaste script. Er zijn twee wijzigingen: het jaar en de Belgische feestdagen voor 2026.
"""

```python
"""
Vakantieoverzicht 2026 - Excel template generator
Vereiste: pip install openpyxl
Uitvoeren: python maak_vakantieoverzicht.py
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import datetime

wb = Workbook()
ws = wb.active
ws.title = "Vakantieoverzicht 2026"

COLOR_HEADER_BG = "1F3864"
COLOR_MONTH_BG = "2E75B6"
COLOR_WEEKEND = "D9D9D9"
COLOR_FEESTDAG = "FFD700"
COLOR_PERSON_BG = ["DCE6F1", "E2EFDA", "FCE4D6", "EAD1DC", "D9EAD3", "FFF2CC"]

thin = Side(style='thin', color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# Belgische feestdagen 2026
feestdagen = {
    datetime.date(2026, 1, 1),   # Nieuwjaar
    datetime.date(2026, 4, 6),   # Paasmaandag
    datetime.date(2026, 5, 1),   # Dag van de Arbeid
    datetime.date(2026, 5, 14),  # OLH Hemelvaart
    datetime.date(2026, 5, 25),  # Pinkstermaandag
    datetime.date(2026, 7, 21),  # Nationale feestdag
    datetime.date(2026, 8, 15),  # OLV Hemelvaart
    datetime.date(2026, 11, 1),  # Allerheiligen
    datetime.date(2026, 11, 11), # Wapenstilstand
    datetime.date(2026, 12, 25), # Kerstmis
}

persons = ["Persoon 1", "Persoon 2", "Persoon 3", "Persoon 4", "Persoon 5", "Persoon 6"]
month_names = ["Januari","Februari","Maart","April","Mei","Juni",
               "Juli","Augustus","September","Oktober","November","December"]

TITLE_ROW, MONTH_ROW, DAY_NUM_ROW, DAY_NAME_ROW, PERSON_START_ROW, NAME_COL = 1, 2, 3, 4, 5, 1

day_cols = {}
col = 2
month_col_ranges = {}

for month in range(1, 13):
    start_col = col
    d = datetime.date(2026, month, 1)
    while d.month == month:
        day_cols[d] = col
        col += 1
        d += datetime.timedelta(days=1)
    month_col_ranges[month] = (start_col, col - 1)

total_cols = col - 1

ws.row_dimensions[TITLE_ROW].height = 28
t = ws.cell(row=TITLE_ROW, column=1, value="VAKANTIEOVERZICHT 2026")
t.font = Font(name="Arial", bold=True, size=14, color="FFFFFF")
t.fill = PatternFill("solid", start_color=COLOR_HEADER_BG)
t.alignment = Alignment(horizontal="center", vertical="center")
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
ws.column_dimensions['A'].width = 14

ws.row_dimensions[MONTH_ROW].height = 18
nc = ws.cell(row=2, column=1, value="Naam")
nc.font = Font(name="Arial", bold=True, size=9, color="FFFFFF")
nc.fill = PatternFill("solid", start_color=COLOR_HEADER_BG)
nc.alignment = Alignment(horizontal="center", vertical="center")
nc.border = border

for month in range(1, 13):
    sc, ec = month_col_ranges[month]
    cell = ws.cell(row=2, column=sc, value=month_names[month - 1])
    cell.font = Font(name="Arial", bold=True, size=9, color="FFFFFF")
    cell.fill = PatternFill("solid", start_color=COLOR_MONTH_BG)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border
    if ec > sc:
        ws.merge_cells(start_row=2, start_column=sc, end_row=2, end_column=ec)

ws.row_dimensions[DAY_NUM_ROW].height = 14
ws.row_dimensions[DAY_NAME_ROW].height = 14
for r, label in [(3, "#"), (4, "Dag")]:
    c = ws.cell(row=r, column=1, value=label)
    c.font = Font(name="Arial", bold=True, size=8, color="FFFFFF")
    c.fill = PatternFill("solid", start_color=COLOR_HEADER_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = border

day_abbr = ["Ma","Di","Wo","Do","Vr","Za","Zo"]

for d, col_i in day_cols.items():
    ws.column_dimensions[get_column_letter(col_i)].width = 2.8
    is_weekend = d.weekday() >= 5
    is_feest = d in feestdagen

    cn = ws.cell(row=3, column=col_i, value=d.day)
    cn.font = Font(name="Arial", size=7, bold=is_feest)
    cn.alignment = Alignment(horizontal="center", vertical="center")
    cn.border = border
    cn.fill = PatternFill("solid", start_color=COLOR_FEESTDAG if is_feest else (COLOR_WEEKEND if is_weekend else "FFFFFF"))

    cd = ws.cell(row=4, column=col_i, value=day_abbr[d.weekday()])
    cd.font = Font(name="Arial", size=7, italic=True)
    cd.alignment = Alignment(horizontal="center", vertical="center")
    cd.border = border
    cd.fill = PatternFill("solid", start_color=COLOR_FEESTDAG if is_feest else (COLOR_WEEKEND if is_weekend else "FFFFFF"))

for p_idx, person in enumerate(persons):
    row = 5 + p_idx
    ws.row_dimensions[row].height = 14
    bg = COLOR_PERSON_BG[p_idx]
    nc2 = ws.cell(row=row, column=1, value=person)
    nc2.font = Font(name="Arial", bold=True, size=8)
    nc2.fill = PatternFill("solid", start_color=bg)
    nc2.alignment = Alignment(horizontal="left", vertical="center")
    nc2.border = border
    for d, col_i in day_cols.items():
        is_weekend = d.weekday() >= 5
        is_feest = d in feestdagen
        cell = ws.cell(row=row, column=col_i, value="")
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(name="Arial", size=7)
        cell.fill = PatternFill("solid", start_color=COLOR_FEESTDAG if is_feest else (COLOR_WEEKEND if is_weekend else "FFFFFF"))

wb.save("Vakantieoverzicht_2026.xlsx")
print("Klaar: Vakantieoverzicht_2026.xlsx")
```

"""
Gewijzigd ten opzichte van de 2025-versie: alle verwijzingen naar `2025` zijn vervangen door `2026`, en de feestdagen zijn herberekend. Paasmaandag valt in 2026 op 6 april (Pasen 5 april), wat de afgeleide data voor Hemelvaart (14 mei, 39 dagen na Pasen) en Pinkstermaandag (25 mei, 50 dagen na Pasen) bepaalt.
"""